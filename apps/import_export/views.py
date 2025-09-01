import io
import re
import os
from datetime import datetime
from collections import defaultdict
import logging
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.files.base import ContentFile, File
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views import View
from unidecode import unidecode
from django.conf import settings
from .forms import AdminExportByCategoryForm, AdminImportFileForm
from .models import ImportExportLog
from apps.products.models import Product, FilterCategory, FilterValue, ProductImage
from apps.menu.models import MenuCatalog
from apps.gallery.models import GalleryImage

logger = logging.getLogger(__name__)

# 1. ------ VUE EXPORT ----------

@method_decorator(staff_member_required, name='dispatch')
class ProductExportSetupView(View):
    """Affiche le formulaire pour choisir la catégorie à exporter."""
    form_class = AdminExportByCategoryForm
    template_name = 'import-export/product_export_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = { 
            'title': 'Экспорт Продуктов',
            'form': form, 
            'media': form.media
        }
        return render(request, self.template_name, context)



# @method_decorator(staff_member_required, name='dispatch')
# class ProductExportDownloadView(View):
#     """Génère, télécharge le fichier Excel et enregistre un log."""
    
#     def get(self, request, *args, **kwargs):
#         form = AdminExportByCategoryForm(request.GET)
#         if not form.is_valid():
#             messages.error(request, f"Неверные параметры для экспорта: {form.errors.as_text()}")
#             return redirect('import_export:product_export_setup')

#         category = form.cleaned_data['category']
#         user_selected_filters = form.cleaned_data['filter_categories_to_export']
        
#         logger.info(f"User '{request.user}' initiated export for category: '{category.name}'.")

#         products_to_export = Product.objects.filter(
#             category__in=category.get_descendants(include_self=True)
#         ).order_by('title')

#         if not products_to_export.exists():
#             messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
#             return redirect('import_export:product_export_setup')

#         if user_selected_filters.exists():
#             filters_for_export = user_selected_filters.order_by('order', 'name')
#         else:
#             used_filter_ids = set(FilterCategory.objects.filter(values__products__in=products_to_export).values_list('pk', flat=True))
#             filters_for_export = FilterCategory.objects.filter(pk__in=used_filter_ids).order_by('order', 'name')
        
#         products_to_export = products_to_export.prefetch_related(
#             'images',
#             Prefetch('filters', queryset=FilterValue.objects.filter(category__in=filters_for_export).select_related('category'))
#         )
        
#         try:
#             workbook = openpyxl.Workbook()
#             sheet = workbook.active
#             sheet.title = slugify(unidecode(category.name))[:30]

#             headers = [
#                 'ID Продукта', 
#                 'Артикул (SKU)', 
#                 'Название', 
#                 'ID Категории', 
#                 'Базовая цена', 
#                 'Тип цены', 
#                 'Ед. изм.', 
#                 'Описание', 
#                 'Ссылки на изображения (|)'
#             ]
#             filter_slug_map = {f.slug: f.name for f in filters_for_export}
#             for slug, name in sorted(filter_slug_map.items()):
#                 headers.append(f"Фильтр: {name} ({slug})")
#             sheet.append(headers)

#             header_font = Font(bold=True)
#             for col_num, _ in enumerate(headers, 1):
#                 sheet.cell(row=1, column=col_num).font = header_font

#             for product in products_to_export.iterator(chunk_size=2000):
#                 image_urls = "|".join([request.build_absolute_uri(img.image.url) for img in product.images.all()])
#                 row_data = [
#                     product.id, 
#                     product.sku, 
#                     product.title, 
#                     product.category_id,
#                     product.base_price, 
#                     product.get_price_type_display(), 
#                     product.unit,
#                     product.description, 
#                     image_urls,
#                 ]
#                 product_filters = defaultdict(list)
#                 for fv in getattr(product, 'prefetched_filters_for_export', product.filters.all()):
#                     if fv.category.slug in filter_slug_map:
#                         product_filters[fv.category.slug].append(fv.value)
#                 for slug, name in sorted(filter_slug_map.items()):
#                     row_data.append("|".join(sorted(product_filters.get(slug, []))))
#                 sheet.append(row_data)

#             excel_buffer = io.BytesIO()
#             workbook.save(excel_buffer)
#             excel_buffer.seek(0)

#             timestamp = datetime.now().strftime('%Y%m%d_%H%M')
#             filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            
#             ImportExportLog.objects.create(
#                 user=request.user, action=ImportExportLog.ACTION_EXPORT, file_name=filename,
#                 file=ContentFile(excel_buffer.read(), name=filename),
#                 details={
#                     'summary': f"Exported {products_to_export.count()} products from category '{category.name}'.",
#                     'category_id': category.id, 'product_count': products_to_export.count()
#                 }
#             )
#             excel_buffer.seek(0)
            
#             response = HttpResponse(
#                 excel_buffer.read(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
#             excel_buffer.close()
#             return response
#         except Exception as e:
#             logger.error(f"Failed to generate Excel export: {e}", exc_info=True)
#             messages.error(request, f"Произошла ошибка при создании файла Excel: {e}")
#             return redirect('import_export:product_export_setup')


@method_decorator(staff_member_required, name='dispatch')
class ProductExportDownloadView(View):
    """
    Génère et télécharge un fichier Excel pour une catégorie de produits donnée.
    Exporte des champs étendus, y compris les noms de fichiers d'images,
    et enregistre une copie de l'export dans le journal.
    """
    
    def get(self, request, *args, **kwargs):
        form = AdminExportByCategoryForm(request.GET)
        if not form.is_valid():
            messages.error(request, f"Неверные параметры для экспорта: {form.errors.as_text()}")
            return redirect('import_export:product_export_setup')

        category = form.cleaned_data['category']
        user_selected_filters = form.cleaned_data['filter_categories_to_export']
        
        logger.info(f"User '{request.user}' initiated export for category: '{category.name}'.")

        # 1. RÉCUPÉRER LES PRODUITS DE LA CATÉGORIE ET DE SES ENFANTS
        products_to_export = Product.objects.filter(
            category__in=category.get_descendants(include_self=True)
        ).order_by('base_name')

        if not products_to_export.exists():
            messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
            return redirect('import_export:product_export_setup')

        # 2. DÉTERMINER LES COLONNES DE FILTRES À INCLURE
        if user_selected_filters.exists():
            filters_for_export = user_selected_filters.order_by('order', 'name')
        else:
            used_filter_ids = set(FilterCategory.objects.filter(values__products__in=products_to_export).values_list('pk', flat=True))
            filters_for_export = FilterCategory.objects.filter(pk__in=used_filter_ids).order_by('order', 'name')
        
        # 3. PRÉCHARGER EFFICACEMENT TOUTES LES DONNÉES NÉCESSAIRES
        products_to_export = products_to_export.prefetch_related(
            # On précharge à travers deux niveaux de relation pour les images
            Prefetch('images', queryset=ProductImage.objects.select_related('gallery_image').order_by('order')),
            # On précharge les filtres et leurs catégories pour la performance
            Prefetch('filters', queryset=FilterValue.objects.filter(category__in=filters_for_export).select_related('category'))
        )
        
        try:
            # 4. CRÉATION DU FICHIER EXCEL EN MÉMOIRE
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = slugify(unidecode(category.name))[:30]

            # Définition des en-têtes en russe
            headers = [
                'ID Продукта', 'Артикул (SKU)', 'Базовое название', 'Полное название (авто)', 
                'ID Категории', 'Базовая цена', 'Тип цены', 'Ед. изм.', 'Описание', 
                'Имена файлов изображений (|)'
            ]
            filter_slug_map = {f.slug: f.name for f in filters_for_export}
            for slug, name in sorted(filter_slug_map.items()):
                headers.append(f"Фильтр: {name} ({slug})")
            sheet.append(headers)

            # Style pour les en-têtes
            header_font = Font(bold=True)
            for col_num, _ in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).font = header_font

            # 5. REMPLISSAGE DES LIGNES AVEC LES DONNÉES DES PRODUITS
            for product in products_to_export.iterator(chunk_size=500):
                # On utilise os.path.basename pour n'extraire que le nom du fichier
                image_filenames = "|".join([
                    os.path.basename(pi.gallery_image.image.name) 
                    for pi in product.images.all() 
                    if pi.gallery_image and pi.gallery_image.image
                ])
                
                row_data = [
                    product.id,
                    product.sku,
                    product.base_name,
                    product.full_title, # Utilise la propriété dynamique
                    product.category_id,
                    product.base_price,
                    product.get_price_type_display(),
                    product.unit,
                    product.description,
                    image_filenames,
                ]
                
                product_filters = defaultdict(list)
                for fv in product.filters.all():
                    if fv.category.slug in filter_slug_map:
                        product_filters[fv.category.slug].append(fv.value)
                
                for slug, name in sorted(filter_slug_map.items()):
                    row_data.append("|".join(sorted(product_filters.get(slug, []))))
                
                sheet.append(row_data)

            # 6. SAUVEGARDE DU FICHIER EN MÉMOIRE ET DANS LE JOURNAL
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            
            ImportExportLog.objects.create(
                user=request.user,
                action=ImportExportLog.ACTION_EXPORT,
                file_name=filename,
                file=ContentFile(excel_buffer.read(), name=filename),
                details={
                    'summary': f"Exported {products_to_export.count()} products from category '{category.name}'.",
                    'category_id': category.id,
                    'product_count': products_to_export.count()
                }
            )
            excel_buffer.seek(0)
            
            # 7. CRÉATION DE LA RÉPONSE HTTP POUR LE TÉLÉCHARGEMENT
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            excel_buffer.close()
            return response
            
        except Exception as e:
            logger.error(f"Failed to generate Excel export for category '{category.name}': {e}", exc_info=True)
            messages.error(request, f"Произошла ошибка при создании файла Excel: {e}")
            return redirect('import_export:product_export_setup')




# ------- Ajax FilterCategory Selection ------
#  
@staff_member_required
def ajax_get_relevant_filters(request, category_id):
    """
    Retourne une liste d'IDs des FilterCategory qui sont utilisés par les produits
    appartenant DIRECTEMENT à la catégorie donnée.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        # On vérifie juste que la catégorie existe
        if not MenuCatalog.objects.filter(pk=category_id).exists():
            raise MenuCatalog.DoesNotExist
        
        products_in_category = Product.objects.filter(category_id=category_id)
        
        relevant_filter_cat_ids = list(FilterCategory.objects.filter(values__products__in=products_in_category).distinct().values_list('pk', flat=True))
        
        logger.debug(f"AJAX (direct only): Relevant filter category IDs for category {category_id}: {relevant_filter_cat_ids}")
        return JsonResponse({'relevant_filter_ids': relevant_filter_cat_ids})

    except MenuCatalog.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching relevant filters for category {category_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Server error'}, status=500)


# 2. ------ VUE EXPORT ----------

@method_decorator(staff_member_required, name='dispatch')
class ProductImportView(View):
    """Affiche le formulaire pour uploader le fichier Excel."""
    form_class = AdminImportFileForm
    template_name = 'import-export/product_import_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {'title': 'Импорт Продуктов из Excel', 'form': form}
        return render(request, self.template_name, context)




# @method_decorator(staff_member_required, name='dispatch')
# @method_decorator(transaction.atomic, name='dispatch')
# class ProductImportProcessView(View):
#     """
#     Traite un fichier Excel pour créer ou mettre à jour des produits, avec une
#     validation stricte, une journalisation complète et des logs de débogage.
#     """
#     form_class = AdminImportFileForm
#     template_name = 'import-export/product_import_results.html'

#     def post(self, request, *args, **kwargs):
#         form = self.form_class(request.POST, request.FILES)
#         if not form.is_valid():
#             messages.error(request, "Ошибка загрузки файла. Убедитесь, что вы выбрали файл .xlsx.")
#             return render(request, 'import_export/product_import_setup.html', {'form': form}, status=400)

#         excel_file = request.FILES['file']
#         log_entry = ImportExportLog.objects.create(
#             user=request.user,
#             action=ImportExportLog.ACTION_IMPORT,
#             file_name=excel_file.name,
#             file=excel_file
#         )
#         logger.info(f"User '{request.user}' initiated import with file: {excel_file.name}")

#         import_summary = {
#             'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'failed_rows': [],
#             'created_filter_categories': 0, 'created_filter_values': 0, 'images_attached': 0
#         }

#         try:
#             workbook = openpyxl.load_workbook(excel_file, data_only=True)
#             sheet = workbook.active
            
#             # --- NORMALISATION ET MAPPING DES EN-TÊTES ---
#             EXPECTED_HEADERS = {
#                 'id продукта': 'ID Продукта',
#                 'артикул (sku)': 'Артикул (SKU)',
#                 'базовое название': 'Базовое название',
#                 'id категории': 'ID Категории',
#                 'базовая цена': 'Базовая цена',
#                 'тип цены': 'Тип цены',
#                 'ед. изм.': 'Ед. изм.',
#                 'описание': 'Описание',
#                 'имена файлов изображений (|)': 'Имена файлов изображений (|)',
#             }

#             header_map = {}
#             raw_headers_from_excel = [str(cell.value or '').strip() for cell in sheet[1]]
            
#             for i, raw_header in enumerate(raw_headers_from_excel):
#                 normalized_header = raw_header.lower().strip()
#                 official_header_name = EXPECTED_HEADERS.get(normalized_header)
#                 if official_header_name:
#                     header_map[official_header_name] = i
            
#             logger.info(f"Header mapping successful: {header_map}")
            
#             required_columns = ['Базовое название', 'ID Категории']
#             for col in required_columns:
#                 if col not in header_map:
#                     raise ValueError(f"Отсутствует обязательная колонка '{col}' в файле Excel. Найденные колонки: {raw_headers_from_excel}")

#             filter_col_map = {}
#             header_regex = re.compile(r'Фильтр:\s*(?P<name>.+?)\s*\((?P<slug>[\w-]+)\)(?:\s*\[(?P<unit>.*?)\])?')
#             for i, header in enumerate(raw_headers_from_excel):
#                 match = header_regex.match(header)
#                 if match:
#                     data = match.groupdict()
#                     filter_name, filter_slug, filter_unit = data['name'].strip(), data['slug'].strip(), (data.get('unit') or '').strip()
#                     if filter_name and filter_slug:
#                         filter_category, created = FilterCategory.objects.get_or_create(slug=filter_slug, defaults={'name': filter_name, 'unit': filter_unit})
#                         if created: import_summary['created_filter_categories'] += 1
#                         filter_col_map[i] = filter_category
            
#             logger.info(f"Mapped {len(filter_col_map)} filter columns.")

#             # --- TRAITEMENT DES LIGNES ---
#             for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#                 import_summary['processed'] += 1
                
#                 base_name = str(row[header_map['Базовое название']]).strip()
#                 category_id_str = str(row[header_map['ID Категории']]).strip()
#                 sku = str(row[header_map.get('Артикул (SKU)', '')] or '').strip()
                
#                 if not base_name or not category_id_str:
#                     import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Отсутствует Base Name или Category ID'}); continue
#                 try:
#                     target_category = MenuCatalog.objects.get(pk=int(category_id_str))
#                 except (ValueError, TypeError, MenuCatalog.DoesNotExist):
#                     import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': f"Категория с ID '{category_id_str}' не найдена"}); continue

#                 product, created = Product.objects.get_or_create(sku=sku) if sku else (None, True)
#                 if created:
#                     product = Product(sku=sku if sku else None); import_summary['created'] += 1
                
#                 product.base_name, product.category = base_name, target_category
#                 try: product.base_price = Decimal(row[header_map['Базовая цена']])
#                 except (InvalidOperation, TypeError, KeyError): pass
                
#                 product.save()

#                 filters_to_set = []
#                 for col_idx, filter_category in filter_col_map.items():
#                     values_str = str(row[col_idx] or '').strip()
#                     if values_str:
#                         values = [v.strip() for v in values_str.split('|') if v.strip()]
#                         for value_name in values:
#                             fv, fv_created = FilterValue.objects.get_or_create(category=filter_category, value=value_name)
#                             if fv_created: import_summary['created_filter_values'] += 1
#                             filters_to_set.append(fv)
#                 product.filters.set(filters_to_set)
                
#                 if 'Имена файлов изображений (|)' in header_map:
#                     filenames_str = str(row[header_map['Имена файлов изображений (|)']] or '').strip()
#                     if filenames_str:
#                         product.images.all().delete()
#                         filenames = [fname.strip() for fname in filenames_str.split('|') if fname.strip()]
#                         for i, filename in enumerate(filenames):
#                             gallery_image = GalleryImage.objects.filter(image__endswith=filename).first()
#                             if gallery_image:
#                                 ProductImage.objects.create(product=product, gallery_image=gallery_image, order=i, is_main=(i == 0))
#                                 import_summary['images_attached'] += 1
#                             else:
#                                 logger.warning(f"Row {row_idx}: Image filename '{filename}' not found in Gallery.")

#                 product.save()
#                 product.category.applicable_filters.add(*filter_col_map.values())
#                 if not created: import_summary['updated'] += 1
            
#             summary_message = (f"Импорт завершен. Обработано: {import_summary['processed']}, "
#                             f"Создано: {import_summary['created']}, Обновлено: {import_summary['updated']}.")
#             messages.success(request, summary_message)
#             log_entry.details = import_summary
#             log_entry.save(update_fields=['details'])
        
#         except ValueError as e:
#             messages.error(request, str(e)); log_entry.details = {'summary': 'Import failed: Validation Error.', 'error': str(e)}; log_entry.save()
#             return redirect('import_export:product_import')
#         except Exception as e:
#             error_message = f"Произошла критическая ошибка: {e}"; logger.error(error_message, exc_info=True); messages.error(request, error_message)
#             log_entry.details = {'summary': 'Import failed critically.', 'error': str(e)}; log_entry.save()
#             return redirect('import_export:product_import')

#         context = {'title': 'Результаты Импорта', 'summary': import_summary}
#         return render(request, self.template_name, context)


@method_decorator(staff_member_required, name='dispatch')
@method_decorator(transaction.atomic, name='dispatch')
class ProductImportProcessView(View):
    """
    Traite un fichier Excel pour créer ou mettre à jour des produits.
    Gère la création/mise à jour, l'attachement d'images, la génération
    de filtres, et une journalisation complète.
    """
    form_class = AdminImportFileForm
    template_name = 'import-export/product_import_results.html'

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Ошибка загрузки файла. Убедитесь, что вы выбрали файл .xlsx.")
            return render(request, 'import_export/product_import_setup.html', {'form': form}, status=400)

        excel_file = request.FILES['file']
        log_entry = ImportExportLog.objects.create(
            user=request.user, action=ImportExportLog.ACTION_IMPORT,
            file_name=excel_file.name, file=excel_file
        )
        logger.info(f"User '{request.user}' initiated import with file: {excel_file.name}")

        import_summary = {
            'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'failed_rows': [],
            'created_filter_categories': 0, 'created_filter_values': 0, 'images_attached': 0
        }

        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            headers_excel = [str(cell.value or '').strip() for cell in sheet[1]]
            
            header_map = {h: i for i, h in enumerate(headers_excel)}
            
            required_columns = ['Базовое название', 'ID Категории']
            for col in required_columns:
                if col not in header_map:
                    raise ValueError(f"Отсутствует обязательная колонка '{col}' в файле Excel.")

            filter_col_map = {}
            header_regex = re.compile(r'Фильтр:\s*(?P<name>.+?)\s*\((?P<slug>[\w-]+)\)(?:\s*\[(?P<unit>.*?)\])?')
            for i, header in enumerate(headers_excel):
                match = header_regex.match(header)
                if match:
                    data = match.groupdict()
                    filter_name, filter_slug, filter_unit = data['name'].strip(), data['slug'].strip(), (data.get('unit') or '').strip()
                    if filter_name and filter_slug:
                        filter_category, created = FilterCategory.objects.get_or_create(
                            slug=filter_slug, defaults={'name': filter_name, 'unit': filter_unit}
                        )
                        if created: import_summary['created_filter_categories'] += 1
                        filter_col_map[i] = filter_category

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                import_summary['processed'] += 1
                
                sku = str(row[header_map.get('Артикул (SKU)', '')] or '').strip()
                base_name = str(row[header_map['Базовое название']]).strip()
                category_id_str = str(row[header_map['ID Категории']]).strip()
                
                if not base_name or not category_id_str:
                    import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Отсутствует Base Name или Category ID'}); continue

                try:
                    target_category = MenuCatalog.objects.get(pk=int(category_id_str))
                except (ValueError, TypeError, MenuCatalog.DoesNotExist):
                    import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': f"Категория с ID '{category_id_str}' не найдена"}); continue

                product, created = Product.objects.get_or_create(sku=sku) if sku else (None, True)
                if created:
                    product = Product(sku=sku if sku else None); import_summary['created'] += 1
                
                product.base_name = base_name; product.category = target_category
                try: product.base_price = Decimal(row[header_map['Базовая цена']])
                except (InvalidOperation, TypeError, KeyError): pass
                
                if 'Тип цены' in header_map and row[header_map['Тип цены']]:
                    price_type_val = str(row[header_map['Тип цены']]).strip()
                    type_key = next((k for k, v in Product.PRICE_TYPE_CHOICES if v == price_type_val), None)
                    if type_key: product.price_type = type_key

                product.save() # Première sauvegarde: génère PK, SKU de base, slug de base, titre de base

                filters_to_set = []
                for col_idx, filter_category in filter_col_map.items():
                    values_str = str(row[col_idx] or '').strip()
                    if values_str:
                        values = [v.strip() for v in values_str.split('|') if v.strip()]
                        for value_name in values:
                            fv, fv_created = FilterValue.objects.get_or_create(category=filter_category, value=value_name)
                            if fv_created: import_summary['created_filter_values'] += 1
                            filters_to_set.append(fv)
                product.filters.set(filters_to_set)
                
                if 'Имена файлов изображений (|)' in header_map:
                    filenames_str = str(row[header_map['Имена файлов изображений (|)']]).strip()
                    if filenames_str:
                        product.images.all().delete()
                        filenames = [fname.strip() for fname in filenames_str.split('|') if fname.strip()]
                        for i, filename in enumerate(filenames):
                            gallery_image = GalleryImage.objects.filter(image__endswith=filename).first()
                            if gallery_image:
                                ProductImage.objects.create(product=product, gallery_image=gallery_image, order=i, is_main=(i == 0))
                                import_summary['images_attached'] += 1
                            else:
                                logger.warning(f"Row {row_idx}: Image filename '{filename}' not found in Gallery.")
                
                # Sauvegarde finale pour régénérer le titre et le slug avec les filtres
                product.save()
                
                product.category.applicable_filters.add(*filter_col_map.values())
                if not created: import_summary['updated'] += 1

            summary_message = (f"Импорт завершен. Обработано: {import_summary['processed']}, "
                            f"Создано: {import_summary['created']}, Обновлено: {import_summary['updated']}.")
            messages.success(request, summary_message)
            log_entry.details = import_summary
            log_entry.save(update_fields=['details'])
        
        except ValueError as e:
            messages.error(request, str(e)); log_entry.details = {'summary': 'Import failed: Validation Error.', 'error': str(e)}; log_entry.save()
            return redirect('import_export:product_import')
        except Exception as e:
            error_message = f"Произошла критическая ошибка: {e}"; logger.error(error_message, exc_info=True); messages.error(request, error_message)
            log_entry.details = {'summary': 'Import failed critically.', 'error': str(e)}; log_entry.save()
            return redirect('import_export:product_import')

        context = {'title': 'Результаты Импорта', 'summary': import_summary}
        return render(request, self.template_name, context)