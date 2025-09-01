import io
import re
from datetime import datetime
from collections import defaultdict
import logging
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views import View
from unidecode import unidecode

from .forms import AdminExportByCategoryForm, AdminImportFileForm
from .models import ImportExportLog
from apps.products.models import Product, FilterCategory, FilterValue
from apps.menu.models import MenuCatalog

logger = logging.getLogger(__name__)

# ===================================================================
# VUE DE CONFIGURATION POUR L'EXPORT
# ===================================================================

@method_decorator(staff_member_required, name='dispatch')
class ProductExportSetupView(View):
    """Affiche le formulaire pour choisir la catégorie à exporter."""
    form_class = AdminExportByCategoryForm
    template_name = 'import-export/product_export_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = { 
            'title': 'Экспорт Продуктов',
            'form': form, 
            'media': form.media
        }
        return render(request, self.template_name, context)

# ===================================================================
# VUE DE TRAITEMENT DE L'EXPORT
# ===================================================================

# @method_decorator(staff_member_required, name='dispatch')
# class ProductExportDownloadView(View):
#     """Génère, télécharge le fichier Excel et enregistre un log."""
    
#     def get(self, request, *args, **kwargs):
#         form = AdminExportByCategoryForm(request.GET)
#         if not form.is_valid():
#             messages.error(request, f"Неверные параметры для экспорта: {form.errors.as_text()}")
#             return redirect('import-export:product_export_setup')

#         category = form.cleaned_data['category']
#         user_selected_filters = form.cleaned_data['filter_categories_to_export']
        
#         logger.info(f"User '{request.user}' initiated export for category: '{category.name}'.")

#         products_to_export = Product.objects.filter(
#             category__in=category.get_descendants(include_self=True)
#         ).order_by('title')

#         if not products_to_export.exists():
#             messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
#             return redirect('import_export:product_export_setup')

#         used_filter_ids = set(FilterCategory.objects.filter(values__products__in=products_to_export).values_list('pk', flat=True))
#         user_selected_ids = set(user_selected_filters.values_list('pk', flat=True))
#         final_filter_ids = used_filter_ids.union(user_selected_ids)
        
#         filters_for_export = FilterCategory.objects.filter(pk__in=final_filter_ids).order_by('order', 'name')
        
#         products_to_export = products_to_export.prefetch_related(
#             'images',
#             Prefetch('filters', queryset=FilterValue.objects.filter(category__in=filters_for_export).select_related('category'))
#         )
        
#         try:
#             workbook = openpyxl.Workbook()
#             sheet = workbook.active
#             sheet.title = slugify(unidecode(category.name))[:30]

#             headers = ['SKU', 'Title', 'Category ID', 'Price', 'Price Type', 'Unit', 'Description', 'Image URLs (| separated)']
#             filter_slug_map = {f.slug: f.name for f in filters_for_export}
#             for slug, name in sorted(filter_slug_map.items()):
#                 headers.append(f"Фильтр: {name} ({slug})")
#             sheet.append(headers)

#             for product in products_to_export.iterator(chunk_size=2000):
#                 image_urls = "|".join([request.build_absolute_uri(img.image.url) for img in product.images.all()])
#                 row_data = [
#                     product.sku, product.title, product.category_id,
#                     product.base_price, product.get_price_type_display(), product.unit,
#                     product.description, image_urls,
#                 ]
#                 product_filters = defaultdict(list)
#                 for fv in product.filters.all():
#                     if fv.category.slug in filter_slug_map:
#                         product_filters[fv.category.slug].append(fv.value)
#                 for slug, name in sorted(filter_slug_map.items()):
#                     row_data.append("|".join(sorted(product_filters.get(slug, []))))
#                 sheet.append(row_data)

#             excel_buffer = io.BytesIO()
#             workbook.save(excel_buffer)
#             excel_buffer.seek(0)

#             timestamp = datetime.now().strftime('%Y%m%d_%H%M')
#             filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            
#             ImportExportLog.objects.create(
#                 user=request.user, action=ImportExportLog.ACTION_EXPORT, file_name=filename,
#                 file=ContentFile(excel_buffer.read(), name=filename),
#                 details={
#                     'summary': f"Exported {products_to_export.count()} products from category '{category.name}'.",
#                     'category_id': category.id, 'product_count': products_to_export.count()
#                 }
#             )
#             excel_buffer.seek(0)
            
#             response = HttpResponse(
#                 excel_buffer.read(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
#             excel_buffer.close()
#             return response
#         except Exception as e:
#             logger.error(f"Failed to generate Excel export: {e}", exc_info=True)
#             messages.error(request, f"Произошла ошибка при создании файла Excel: {e}")
#             return redirect('import_export:product_export_setup')




@method_decorator(staff_member_required, name='dispatch')
class ProductExportDownloadView(View):
    """Génère, télécharge le fichier Excel et enregistre un log."""
    
    def get(self, request, *args, **kwargs):
        form = AdminExportByCategoryForm(request.GET)
        if not form.is_valid():
            messages.error(request, f"Неверные параметры для экспорта: {form.errors.as_text()}")
            return redirect('import_export:product_export_setup')

        category = form.cleaned_data['category']
        user_selected_filters = form.cleaned_data['filter_categories_to_export']
        
        logger.info(f"User '{request.user}' initiated export for category: '{category.name}'.")

        products_to_export = Product.objects.filter(
            category__in=category.get_descendants(include_self=True)
        ).order_by('title')

        if not products_to_export.exists():
            messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
            return redirect('import_export:product_export_setup')

        if user_selected_filters.exists():
            filters_for_export = user_selected_filters.order_by('order', 'name')
        else:
            used_filter_ids = set(FilterCategory.objects.filter(values__products__in=products_to_export).values_list('pk', flat=True))
            filters_for_export = FilterCategory.objects.filter(pk__in=used_filter_ids).order_by('order', 'name')
        
        products_to_export = products_to_export.prefetch_related(
            'images',
            Prefetch('filters', queryset=FilterValue.objects.filter(category__in=filters_for_export).select_related('category'))
        )
        
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = slugify(unidecode(category.name))[:30]

            headers = [
                'ID Продукта', 
                'Артикул (SKU)', 
                'Название', 
                'ID Категории', 
                'Базовая цена', 
                'Тип цены', 
                'Ед. изм.', 
                'Описание', 
                'Ссылки на изображения (|)'
            ]
            filter_slug_map = {f.slug: f.name for f in filters_for_export}
            for slug, name in sorted(filter_slug_map.items()):
                headers.append(f"Фильтр: {name} ({slug})")
            sheet.append(headers)

            header_font = Font(bold=True)
            for col_num, _ in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).font = header_font

            for product in products_to_export.iterator(chunk_size=2000):
                image_urls = "|".join([request.build_absolute_uri(img.image.url) for img in product.images.all()])
                row_data = [
                    product.id, 
                    product.sku, 
                    product.title, 
                    product.category_id,
                    product.base_price, 
                    product.get_price_type_display(), 
                    product.unit,
                    product.description, 
                    image_urls,
                ]
                product_filters = defaultdict(list)
                # On utilise l'attribut préchargé pour la performance
                for fv in getattr(product, 'prefetched_filters_for_export', product.filters.all()):
                    if fv.category.slug in filter_slug_map:
                        product_filters[fv.category.slug].append(fv.value)
                for slug, name in sorted(filter_slug_map.items()):
                    row_data.append("|".join(sorted(product_filters.get(slug, []))))
                sheet.append(row_data)

            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            
            ImportExportLog.objects.create(
                user=request.user, action=ImportExportLog.ACTION_EXPORT, file_name=filename,
                file=ContentFile(excel_buffer.read(), name=filename),
                details={
                    'summary': f"Exported {products_to_export.count()} products from category '{category.name}'.",
                    'category_id': category.id, 'product_count': products_to_export.count()
                }
            )
            excel_buffer.seek(0)
            
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            excel_buffer.close()
            return response
        except Exception as e:
            logger.error(f"Failed to generate Excel export: {e}", exc_info=True)
            messages.error(request, f"Произошла ошибка при создании файла Excel: {e}")
            return redirect('import_export:product_export_setup')

# ===================================================================
# VUES D'IMPORTATION
# ===================================================================

@method_decorator(staff_member_required, name='dispatch')
class ProductImportView(View):
    """Affiche le formulaire pour uploader le fichier Excel."""
    form_class = AdminImportFileForm
    template_name = 'import-export/product_import_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {'title': 'Импорт Продуктов из Excel', 'form': form}
        return render(request, self.template_name, context)

@method_decorator(staff_member_required, name='dispatch')
@method_decorator(transaction.atomic, name='dispatch')
class ProductImportProcessView(View):
    """Traite le fichier Excel, met à jour la DB, et enregistre un log."""
    form_class = AdminImportFileForm
    template_name = 'import-export/product_import_results.html'

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Ошибка загрузки файла. Пожалуйста, убедитесь, что вы выбрали файл .xlsx.")
            context = {'title': 'Импорт Продуктов (Ошибка)', 'form': form}
            return render(request, 'import-export/product_import_setup.html', context, status=400)

        excel_file = request.FILES['file']
        log_entry = ImportExportLog.objects.create(
            user=request.user, action=ImportExportLog.ACTION_IMPORT,
            file_name=excel_file.name, file=excel_file
        )
        logger.info(f"User '{request.user}' initiated import with file: {excel_file.name}")

        import_summary = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'failed_rows': []}

        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            headers_excel = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            header_map = {h: i for i, h in enumerate(headers_excel)}
            if 'SKU' not in header_map or 'Title' not in header_map:
                raise ValueError("Отсутствуют обязательные колонки 'SKU' или 'Title'.")

            filter_col_map = {}
            active_cats_by_slug = {fc.slug: fc for fc in FilterCategory.objects.all()}
            slug_regex = re.compile(r'\((\w+[-_]*\w+)\)')
            for i, header in enumerate(headers_excel):
                if header.startswith('Фильтр:'):
                    match = slug_regex.search(header)
                    if match and match.group(1) in active_cats_by_slug:
                        filter_col_map[i] = active_cats_by_slug[match.group(1)]

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                import_summary['processed'] += 1
                sku = str(row[header_map['SKU']]).strip()
                if not sku:
                    import_summary['skipped'] += 1
                    import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Missing SKU'})
                    continue

                product, created = Product.objects.get_or_create(sku=sku, defaults={'title': str(row[header_map['Title']])})
                
                if created:
                    import_summary['created'] += 1
                else:
                    product.title = str(row[header_map['Title']]).strip()
                
                try: product.base_price = Decimal(row[header_map['Price']])
                except (InvalidOperation, TypeError, KeyError): pass
                
                if 'Price Type' in header_map:
                    price_type_val = str(row[header_map['Price Type']]).strip()
                    if price_type_val in dict(Product.PRICE_TYPE_CHOICES).values():
                        # Find the key for the given value
                        type_key = next((k for k, v in Product.PRICE_TYPE_CHOICES if v == price_type_val), None)
                        if type_key: product.price_type = type_key
                
                if 'Unit' in header_map: product.unit = str(row[header_map.get('Unit', '')] or '').strip()
                if 'Description' in header_map: product.description = str(row[header_map.get('Description', '')] or '').strip()
                
                if 'Category ID' in header_map:
                    try:
                        cat_id = int(row[header_map['Category ID']])
                        product.category = MenuCatalog.objects.get(pk=cat_id)
                    except (ValueError, TypeError, MenuCatalog.DoesNotExist): pass
                
                product.save()

                filters_to_set = []
                for col_idx, filter_category in filter_col_map.items():
                    values_str = str(row[col_idx] or '').strip()
                    if values_str:
                        values = [v.strip() for v in values_str.split('|') if v.strip()]
                        for value_name in values:
                            fv, _ = FilterValue.objects.get_or_create(category=filter_category, value=value_name)
                            filters_to_set.append(fv)
                
                product.filters.set(filters_to_set)
                
                if not created:
                    import_summary['updated'] += 1
            
            summary_message = f"Импорт завершен. Обработано: {import_summary['processed']}, Создано: {import_summary['created']}, Обновлено: {import_summary['updated']}, Пропущено: {import_summary['skipped']}."
            messages.success(request, summary_message)
            log_entry.details = {'summary': summary_message, 'failed_rows': import_summary['failed_rows']}
            log_entry.save(update_fields=['details'])
        
        except Exception as e:
            error_message = f"Произошла критическая ошибка при обработке файла: {e}"
            logger.error(error_message, exc_info=True)
            messages.error(request, error_message)
            log_entry.details = {'summary': 'Import failed critically.', 'error': str(e)}
            log_entry.save(update_fields=['details'])
            return redirect('import_export:product_import')

        context = {'title': 'Результаты Импорта', 'summary': import_summary}
        return render(request, self.template_name, context)
    


# @staff_member_required
# def ajax_get_relevant_filters(request, category_id):
#     """
#     Retourne une liste d'IDs des FilterCategory qui sont utilisés
#     par les produits de la catégorie donnée (et de ses descendantes).
#     """
#     if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
#         return JsonResponse({'error': 'AJAX request required'}, status=400)
    
#     try:
#         category = MenuCatalog.objects.get(pk=category_id)
#         # On récupère les produits de la catégorie et de ses descendantes
#         products_in_category = Product.objects.filter(
#             category__in=category.get_descendants(include_self=True)
#         )
        
#         # On trouve les IDs des catégories de filtres utilisées par ces produits
#         relevant_filter_cat_ids = list(
#             FilterCategory.objects.filter(
#                 values__products__in=products_in_category
#             ).distinct().values_list('pk', flat=True)
#         )
        
#         return JsonResponse({'relevant_filter_ids': relevant_filter_cat_ids})

#     except MenuCatalog.DoesNotExist:
#         return JsonResponse({'error': 'Category not found'}, status=404)
#     except Exception as e:
#         logger.error(f"Error fetching relevant filters for category {category_id}: {e}", exc_info=True)
#         return JsonResponse({'error': 'Server error'}, status=500)



@staff_member_required
def ajax_get_relevant_filters(request, category_id):
    """
    Retourne une liste d'IDs des FilterCategory qui sont utilisés par les produits
    appartenant DIRECTEMENT à la catégorie donnée (SANS inclure les descendantes).
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        # On vérifie juste que la catégorie existe
        if not MenuCatalog.objects.filter(pk=category_id).exists():
            raise MenuCatalog.DoesNotExist
        
        # On ne prend que les produits dont la 'category_id' est exactement celle fournie.
        products_in_category = Product.objects.filter(category_id=category_id)
        
        # La suite de la logique reste la même
        relevant_filter_cat_ids = list(FilterCategory.objects.filter(values__products__in=products_in_category).distinct().values_list('pk', flat=True))
        
        logger.debug(f"AJAX (direct only): Relevant filter category IDs for category {category_id}: {relevant_filter_cat_ids}")
        return JsonResponse({'relevant_filter_ids': relevant_filter_cat_ids})

    except MenuCatalog.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching relevant filters for category {category_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Server error'}, status=500)